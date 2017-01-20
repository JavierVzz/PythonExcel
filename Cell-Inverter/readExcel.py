# Javier Vazquez
# Python 3.6

import openpyxl, os, re, pprint, sys, logging

#os.remove("log.txt")
logging.basicConfig(filename="log.txt", level = logging.DEBUG, format =' %(asctime)s - %(levelname)s - %(message)s')

class excelOperations:

    def __init__(self):
        self.__wb = 0
        self.__ws = 0
        self.__excelDict = {}

    @property
    def wb_ws(self):
        return self.__wb, self.__ws

    def displayExcelFiles(self):
        listXlsx = []
        extRegex = re.compile(r".+xlsx$")
        listFiles = os.listdir()
        for file in listFiles:
            ext = extRegex.search(file)
            if ext is not None:
                listXlsx.append(ext.group())
        return listXlsx

    def readExcelFile(self, file):
        if os.path.isfile(file):
            self.__wb = openpyxl.load_workbook(file)
            self.__ws = self.__wb.active
            return self.__wb, self.__ws
        else:
            return False, False

    def displayContent(self):
        for i in range(1, self.__ws.max_row + 1):
            rowToDisplay = ""
            for j in range(1, self.__ws.max_column + 1):
                rowToDisplay += str(self.__ws.cell(row=i, column=j).value) + " "
            print(str(i)+": "+rowToDisplay)
            logging.debug("Row: " + rowToDisplay)

    def createDictionary(self):
        for i in range(2, self.__ws.max_row + 1):
        # for i in range(2, 1000 + 1):
            produce = self.__ws.cell(row=i, column=1).value
            if produce not in self.__excelDict.keys():
                self.__excelDict[produce] = self.__ws.cell(row=i, column=2).value
                logging.debug("Item added to dict- "+str(produce)+": "+str(self.__excelDict[produce]))

    def saveDictionary(self, mode):
        line = ""
        textFile = open("dictionary.txt", mode)
        # sortedDictList = sorted(self.__excelDict.items(), reverse = False, key=lambda x: x[0])
        for key, value in self.__excelDict.items():
            line += str(key) + ": " + str(value) + "\n"
        textFile.write(line)
        textFile.close()

    def loadDictionary(self):
        self.__excelDict = {}
        produceRegex = re.compile(r"^(\w.+):")
        priceRegex = re.compile(r":(.+)$")
        textFile = open("dictionary.txt", "r")
        textLines = textFile.readlines()
        for textLine in textLines:
            matchingProduce = produceRegex.search(textLine)
            matchingPrice = priceRegex.search(textLine)
            if matchingProduce.group(1) not in self.__excelDict.keys():
                self.__excelDict[matchingProduce.group(1)] = float(matchingPrice.group(1))
        textFile.close()

    def printDictionary(self):
        sortedDictList = sorted(self.__excelDict.items(), reverse=False, key=lambda x: x[0])
        for i in range(0, len(sortedDictList), 2):
            print("{0:<20}{1:>5}{2:<5}{3:<20}{4:>5}"
                  .format(sortedDictList[i][0], sortedDictList[i][1], " ",
                          sortedDictList[i+1][0], sortedDictList[i+1][1]))

    def updateKey(self, key, value):
        if key in self.__excelDict.keys():
            self.__excelDict[key] = value
            self.saveDictionary("w")
            return True
        else:
            return False

    def updateExcel(self, key, value):
        for i in range(2, self.__ws.max_row + 1):
        # for i in range(2, 1000 + 1):
            if self.__ws.cell(row=i, column=1).value == key:
                self.__ws.cell(row=i, column=2).value = float(value)
                self.__wb.save("produceSales.xlsx")

if __name__ == '__main__':
    print("Direct access to "+ os.path.basename(__file__))
else:
    print(os.path.basename(__file__)+" class instance")
