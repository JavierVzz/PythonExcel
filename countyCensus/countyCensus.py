# Javier Vazquez

import openpyxl, os, re, pprint, sys, logging
from openpyxl.styles import Font
os.remove("log.txt")
os.remove("countyCensus.txt")
os.remove("countyCensus.xlsx")
logging.basicConfig(filename="log.txt", level = logging.DEBUG, format =' %(asctime)s - %(levelname)s - %(message)s')


class Counter():

    def __init__(self):
        self.__wb = 0
        self.__ws = 0
        self.__wbWrite = openpyxl.Workbook()
        self.__wsWrite = self.__wbWrite.active
        self.__wsWrite.column_dimensions["A"].width = 20
        self.__wsWrite.column_dimensions["B"].width = 15
        self.__wsWrite.column_dimensions["C"].width = 10
        self.__wsWrite.column_dimensions["D"].width = 15
        self.__wsWrite.column_dimensions["E"].width = 15
        self.__dictCensusPerCounty ={}

    def displayXlsxFiles(self):
        listXlsx = []
        extRegex = re.compile(r".+xlsx$")
        listFiles = os.listdir()
        for file in listFiles:
            ext = extRegex.search(file)
            if ext is not None:
                listXlsx.append(ext.group())
        return listXlsx

    def readExcelFile(self, file):
        if os.path.isfile(file):
            self.__wb = openpyxl.load_workbook(file, read_only=True)
            self.__ws = self.__wb.active
            return self.__wb, self.__ws
        else:
            return False, False

    def writeExcelFile(self, id, county, state, tracts, pop):
        self.__wsWrite.append([id, county, state, tracts, pop])
        self.__wbWrite.save("countyCensus.xlsx")

    def writeTextFile(self, countyRow):
        textFile = open("countyCensus.txt", "a")
        textFile.write(countyRow)
        textFile.close()

    def countPerCounty(self):
        lastKey= ""
        for i in range(2, self.__ws.max_row+1):
        #for i in range(self.__ws.max_row-100, self.__ws.max_row+1):
        #for i in range(2, 1349):
            logging.debug("Iteration: " + str(i))
            county = self.__ws.cell(row=i, column=3).value
            state = self.__ws.cell(row=i, column=2).value
            pop = self.__ws.cell(row=i, column=4).value
            idKey = county +"_"+ state
            logging.debug(str(idKey) +" "+ str(county) +" "+ str(state) +" "+ str(pop))
            #  i = 1 are the headers.
            #  i = 2 the values start
            if i == 2:
            #if i == self.__ws.max_row-100:
                self.__dictCensusPerCounty[idKey] = {"County": county, "State": state, "C. Tract": 1, "Population": pop}
                logging.debug("First element added to the dict: " + str(self.__dictCensusPerCounty[idKey]))
                lastKey = idKey
                self.writeTextFile("{0:<30}{1:>20}{2:>10}{3:>15}{4:>15}".format(
                    "ID",
                    "County",
                    "State",
                    "# of Tracts",
                    "Population\n"))
                self.writeExcelFile("ID", "County", "State", "# of Tracts", "Population\n")
                headers = Font(bold=True)
                for j in range(1,6):
                    self.__wsWrite.cell(row = 1, column = j).font = headers

            elif idKey in self.__dictCensusPerCounty:
                self.__dictCensusPerCounty[idKey]["C. Tract"] += 1
                self.__dictCensusPerCounty[idKey]["Population"] += pop
                logging.debug("Added to the dict: " + str(self.__dictCensusPerCounty[idKey]))
                lastKey = idKey

                if i == self.__ws.max_row:
                    print("THE END")
                    logging.debug("**LAST OCCURRENCE " + idKey + "\n, Added to the dict: " + str(self.__dictCensusPerCounty[idKey]))
                    self.writeTextFile("{0:<30}{1:>20}{2:>10}{3:>15}{4:>15}".format(
                        lastKey,
                        str(self.__dictCensusPerCounty[lastKey]["County"]),
                        str(self.__dictCensusPerCounty[lastKey]["State"]),
                        str(self.__dictCensusPerCounty[lastKey]["C. Tract"]),
                        str(self.__dictCensusPerCounty[lastKey]["Population"]) + "\n"))

                    self.writeExcelFile(lastKey,
                        str(self.__dictCensusPerCounty[lastKey]["County"]),
                        str(self.__dictCensusPerCounty[lastKey]["State"]),
                        self.__dictCensusPerCounty[lastKey]["C. Tract"],
                        self.__dictCensusPerCounty[lastKey]["Population"])

            elif idKey not in self.__dictCensusPerCounty:
                self.__dictCensusPerCounty[idKey] = {"County": county, "State": state, "C. Tract": 1, "Population": pop}
                logging.debug("Last Key "+ lastKey +" " + str(self.__dictCensusPerCounty[lastKey]))
                logging.debug("1st OCCURRENCE DIFFERENT COUNTY "+ idKey +"\n, Added to the dict: " + str(self.__dictCensusPerCounty[idKey]))
                self.writeTextFile("{0:<30}{1:>20}{2:>10}{3:>15}{4:>15}".format(
                    lastKey,
                    str(self.__dictCensusPerCounty[lastKey]["County"]),
                    str(self.__dictCensusPerCounty[lastKey]["State"]),
                    str(self.__dictCensusPerCounty[lastKey]["C. Tract"]),
                    str(self.__dictCensusPerCounty[lastKey]["Population"])+"\n"))
                self.writeExcelFile(lastKey,
                    str(self.__dictCensusPerCounty[lastKey]["County"]),
                    str(self.__dictCensusPerCounty[lastKey]["State"]),
                    self.__dictCensusPerCounty[lastKey]["C. Tract"],
                    self.__dictCensusPerCounty[lastKey]["Population"])
                lastKey = idKey


        pprint.pprint(self.__dictCensusPerCounty)


def main():
    logging.debug("Start of program")
    census = Counter()
    listXlsx = census.displayXlsxFiles()
    if len(listXlsx) == 0:
        print("No xlsx files in the current folder")
        input("Press any key to exit!!!")
        sys.exit(0)

    pprint.pprint(listXlsx)
    file = input("File: ")
    census.readExcelFile(file)
    census.countPerCounty()
    print("Done")


main()