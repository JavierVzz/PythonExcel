# Javier Vazquez

import openpyxl, os
from openpyxl.styles import Font

class nbyn:

    def __init__(self, n):
        self.__n = int(n)
        self.__wbWrite = openpyxl.Workbook()
        self.__wsWrite = self.__wbWrite.active

    def addingHeader(self):
        headers = Font(bold=True)
        for rowM in range(1, self.__n+2):
            if rowM == 1:
                self.__wsWrite.cell(row= rowM, column= 1).value = ""
            else:
                self.__wsWrite.cell(row= rowM, column= 1).font = headers
                self.__wsWrite.cell(row= rowM, column= 1).value = rowM-1
        for columnN in range(1, self.__n+2):
            if columnN == 1:
                self.__wsWrite.cell(row= 1, column= columnN).value = ""
            else:
                self.__wsWrite.cell(row= 1, column= columnN).font = headers
                self.__wsWrite.cell(row= 1, column= columnN).value = columnN-1
        self.__wbWrite.save("NbyN.xlsx")

    def populatingTable(self):
        for columnN in range(1, self.__n+1):
            for rowM in range(1, self.__n+1):
                    self.__wsWrite.cell(row= rowM+1, column= columnN+1).value = rowM * columnN
        self.__wbWrite.save("NbyN.xlsx")

    def sizeOfTable(self):
        return self.__n

if __name__ == '__main__':
    print("nbyn is being run directly")
else:
    print("nbyn is being accessed")
    os.chdir("C:\\Users\\Javier\\PycharmProjects\\ExtendedWork\\NbyN")


